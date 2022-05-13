from tkinter import *
from tkinter import ttk
from DES import encryptDES,decryptDES
from math import ceil
from random import randint
from tkinter.filedialog import askopenfile,asksaveasfile
import docx
from openpyxl import load_workbook
import io



FontBig = ('Arial',15,'bold')
FontSmall = ('Arial',13,'bold')
BG = '#a7beaf'

window = Tk()

window.title('Nguyễn Danh Hòa - Ứng dụng thuật toán DES và lược đồ chia sẻ bí mật trong đăng kí bỏ phiếu điện tử.')

window.geometry('1600x700')

window.config(bg=BG)

notebook = ttk.Notebook(window)
notebook.pack(pady=0)

frame1 = Frame(notebook, width=1600, height=700, bg='#a7beaf', padx=20, pady=20)
frame2 = Frame(notebook, width=1600, height=700, bg='#a7beaf', padx=20, pady=20)

frame1.pack(fill='both', expand=1)
frame2.pack(fill='both', expand=1)

notebook.add(frame1, text='Đầu gửi')
notebook.add(frame2, text='Đầu nhận')

label0 = Label(frame1, text='Nhập các giá trị:', font=FontBig, bg=BG)
label0.place(x=0, y=0)

p_label = Label(frame1, text='p:', font=FontSmall, bg=BG)
p_label.place(x=10, y=30)

a1_label = Label(frame1, text='a1:', font=FontSmall, bg=BG)
a1_label.place(x=120, y=30)

a2_label = Label(frame1, text='a2:', font=FontSmall, bg=BG)
a2_label.place(x=230, y=30)

v1_label = Label(frame1, text='Thành viên 1:', font=FontSmall, bg=BG)
v1_label.place(x=340, y=30)

v2_label = Label(frame1, text='Thành viên 2:', font=FontSmall, bg=BG)
v2_label.place(x=540, y=30)

v3_label = Label(frame1, text='Thành viên 3:', font=FontSmall, bg=BG)
v3_label.place(x=740, y=30)

v4_label = Label(frame1, text='Thành viên 4:', font=FontSmall, bg=BG)
v4_label.place(x=940, y=30)

v5_label = Label(frame1, text='Thành viên 5:', font=FontSmall, bg=BG)
v5_label.place(x=1140, y=30)

key_label = Label(frame1, text='Nhập mật khẩu:', font=FontBig, bg=BG)
key_label.place(x=0, y=80)

source1_label = Label(frame1, text='Nhập bản rõ:', font=FontBig, bg=BG)
source1_label.place(x=0, y=130)

ans1_label = Label(frame1, text='Bản mã:', font=FontBig, bg=BG)
ans1_label.place(x=870, y=130)

choose_mem_label = Label(frame2, text='Chọn 3 thành viên để khôi phục khóa', font=FontBig, bg=BG)
choose_mem_label.place(x=600, y=0)

check_label = Label(frame2, text='', font=FontBig, bg=BG)
check_label.place(x=600, y=100)

key_decrypt_label = Label(frame2, text='Khóa bí mật: ', font=FontBig, bg=BG)
key_decrypt_label.place(x=0, y=130)

source2_label = Label(frame2, text='Nhập bản mã:', font=FontBig, bg=BG)
source2_label.place(x=0, y=170)

ans2_label = Label(frame2, text='Bản rõ:', font=FontBig, bg=BG)
ans2_label.place(x=870, y=170)


p_entry = Entry(frame1, width=8, font=FontSmall)
p_entry.place(x=30, y=30)

a1_entry = Entry(frame1, width=8, font=FontSmall)
a1_entry.place(x=150, y=30)

a2_entry = Entry(frame1, width=8, font=FontSmall)
a2_entry.place(x=260, y=30)

v1_entry = Entry(frame1, width=8, font=FontSmall)
v1_entry.place(x=450, y=30)

v2_entry = Entry(frame1, width=8, font=FontSmall)
v2_entry.place(x=650, y=30)

v3_entry = Entry(frame1, width=8, font=FontSmall)
v3_entry.place(x=850, y=30)

v4_entry = Entry(frame1, width=8, font=FontSmall)
v4_entry.place(x=1050, y=30)

v5_entry = Entry(frame1, width=8, font=FontSmall)
v5_entry.place(x=1250, y=30)

key_entry = Entry(frame1, width=20, font=FontBig)
key_entry.place(x=160, y=80)

source1_entry = Text(frame1, width=85, height=17)
source1_entry.place(x=0, y=170)

ans1_entry = Text(frame1, width=85, height=17)
ans1_entry.place(x=870, y=170)

key_decrypt_entry = Entry(frame2, width=20, font=FontBig)
key_decrypt_entry.place(x=160, y=130)

source2_entry = Text(frame2, width=85, height=17)
source2_entry.place(x=0, y=200)

ans2_entry = Text(frame2, width=85, height=17)
ans2_entry.place(x=870, y=200)

var1 = IntVar()
var2 = IntVar()
var3 = IntVar()
var4 = IntVar()
var5 = IntVar()

check1 = Checkbutton(frame2, text='Thành viên 1', variable=var1, font=('Arial', 13, 'bold'), bg='#a7beaf')
check1.place(x=0, y=50)

check2 = Checkbutton(frame2, text='Thành viên 2', variable=var2, font=('Arial', 13, 'bold'), bg='#a7beaf')
check2.place(x=300, y=50)

check3 = Checkbutton(frame2, text='Thành viên 3', variable=var3, font=('Arial', 13, 'bold'), bg='#a7beaf')
check3.place(x=600, y=50)

check4 = Checkbutton(frame2, text='Thành viên 4', variable=var4, font=('Arial', 13, 'bold'), bg='#a7beaf')
check4.place(x=900, y=50)

check5 = Checkbutton(frame2, text='Thành viên 5', variable=var5, font=('Arial', 13, 'bold'), bg='#a7beaf')
check5.place(x=1200, y=50)


def randommem():
    p = randint(50,100)
    a1 = randint(1,p)
    a2 = randint(1,p)
    v1 = randint(10,200)
    v2 = randint(10,200)
    v3 = randint(10,200)
    v4 = randint(10,200)
    v5 = randint(10,200)
    p_entry.delete(0,'end')
    a1_entry.delete(0,'end')
    a2_entry.delete(0,'end')
    v1_entry.delete(0,'end')
    v2_entry.delete(0,'end')
    v3_entry.delete(0,'end')
    v4_entry.delete(0,'end')
    v5_entry.delete(0,'end')
    p_entry.insert(0,str(p))
    a1_entry.insert(0,str(a1))
    a2_entry.insert(0,str(a2))
    v1_entry.insert(0,str(v1))
    v2_entry.insert(0,str(v2))
    v3_entry.insert(0,str(v3))
    v4_entry.insert(0,str(v4))
    v5_entry.insert(0,str(v5))


def randomkey():
    key = randint(10000000,99999999)
    key_entry.delete(0,'end')
    key_entry.insert(0,str(key))


def readWordfile(filename):
    doc = docx.Document(filename)
    result = [p.text for p in doc.paragraphs]
    return '\n'.join(result)


def readExcelfile(filename):
    book = load_workbook(filename)
    sheet = book.active
    ans = ''
    for i in range(1,sheet.max_row+1):
        for j in range(1,sheet.max_column+1):
            ans+= str(sheet.cell(row=i, column=j).value)
            ans+=' '
        ans+='\n'
    return ans

def readTextfile(filename):
    s= ''
    file = io.open(filename, mode='r', encoding='utf-8')
    read = file.readlines()
    for line in read:
        s+=line
    return s

filetypes = [('Word Files', '*.docx'), ('Text Files', '*.txt'), ('Excel Files', '*.xlsx')]

def open():
    global data
    file = askopenfile(mode='r', initialdir='C:/Users/Admin/Desktop/An Toan Va Bao Mat Thong Tin/BTL', title='Select File', filetypes=filetypes)
    s = str(file)
    a = s.find('name')
    b = s.find('mode')
    filename = s[a+6:b-2]
    if 'docx' in filename:
        data = readWordfile(filename)
    elif 'xlsx' in filename:
        data = readExcelfile(filename)
    elif 'txt' in filename:
        data = readTextfile(filename)
    source1_entry.delete(1.0,END)
    source1_entry.insert(INSERT,data)

def decrypt_open():
    global data
    file = askopenfile(mode='r', initialdir='C:/Users/Admin/Desktop/An Toan Va Bao Mat Thong Tin/BTL',
                       title='Select File', filetypes=filetypes)
    s = str(file)
    a = s.find('name')
    b = s.find('mode')
    filename = s[a + 6:b - 2]
    if 'docx' in filename:
        data = readWordfile(filename)
    elif 'xlsx' in filename:
        data = readExcelfile(filename)
    elif 'txt' in filename:
        data = readTextfile(filename)
    source2_entry.delete(1.0, END)
    source2_entry.insert(INSERT, data)


def save():
    f = asksaveasfile(mode='w', defaultextension=".txt")
    if f is None:
        return
    text2save = str(ans1_entry.get(1.0, END))
    f.write(text2save)
    f.close()


def decrypt_save():
    f = asksaveasfile(mode='wb', defaultextension=".txt")
    if f is None:
        return
    text2save = str(ans2_entry.get(1.0, END)).encode('utf8')
    f.write(text2save)
    f.close()


def share_key():
    global m1,m2,m3,m4,m5
    v1 = int(v1_entry.get())
    v2 = int(v2_entry.get())
    v3 = int(v3_entry.get())
    v4 = int(v4_entry.get())
    v5 = int(v5_entry.get())
    a1 = int(a1_entry.get())
    a2 = int(a2_entry.get())
    k = int(key_entry.get())
    m1 = a2 * (v1 ** 2) + a1 * v1 + k
    m2 = a2 * (v2 ** 2) + a1 * v2 + k
    m3 = a2 * (v3 ** 2) + a1 * v3 + k
    m4 = a2 * (v4 ** 2) + a1 * v4 + k
    m5 = a2 * (v5 ** 2) + a1 * v5 + k

def restore_key():
    k = 0
    lstcheckbox = list((var1.get(), var2.get(), var3.get(), var4.get(), var5.get()))
    if not m1:
        check_label.config(text='Chưa chia sẻ khóa', fg='red')
    elif lstcheckbox.count(1) < 3:
        check_label.config(text='Phải chọn ít nhất là 3 thành viên để khôi phục khóa', fg='red')
    else:
        lst_manh = []
        lst_b = []
        lst_v = []
        lst_index = []
        for i in range(len(lstcheckbox)):
            if lstcheckbox[i] == 1:
                lst_index.append(i + 1)
        for i in lst_index:
            if i == 1:
                lst_manh.append(m1)
                lst_v.append(int(v1_entry.get()))
            elif i == 2:
                lst_manh.append(m2)
                lst_v.append(int(v2_entry.get()))
            elif i == 3:
                lst_manh.append(m3)
                lst_v.append(int(v3_entry.get()))
            elif i == 4:
                lst_manh.append(m4)
                lst_v.append(int(v4_entry.get()))
            else:
                lst_manh.append(m5)
                lst_v.append(int(v5_entry.get()))
        for i in range(len(lst_v)):
            lst_temp = lst_v.copy()
            del lst_temp[i]
            b = 1
            for j in lst_temp:
                b *= (j / (j - lst_v[i]))
            lst_b.append(b)
        for i in range(len(lst_manh)):
            k += lst_b[i] * lst_manh[i]
        k = ceil(k)
        check_label.config(text='Khôi phục khóa thành công', fg='green')
        key_decrypt_entry.delete(0,'end')
        key_decrypt_entry.insert(0,str(k))



def encrypt():
    key = key_entry.get()
    text = source1_entry.get(1.0, END)
    ans = encryptDES(key, text)
    ans1_entry.delete(1.0, END)
    ans1_entry.insert(INSERT, ans)


def decrypt():
    key = key_decrypt_entry.get()
    text = source2_entry.get(1.0, END)
    ans = decryptDES(key, text)
    ans2_entry.delete(1.0, END)
    ans2_entry.insert(INSERT, ans)


def exit():
    window.destroy()


rd_btn = Button(frame1, text='Tạo ngẫu nhiên', width=15, height=1, font=('Arial', 10, 'bold'), bg=BG, command=randommem)
rd_btn.place(x=1400, y=30)

rd_key_btn = Button(frame1, text='Tạo ngẫu nhiên', width=15, height=1, font=('Arial', 10, 'bold'), bg=BG, command=randomkey)
rd_key_btn.place(x=400, y=80)

upfile1_btn = Button(frame1, text='Đọc từ file', width=15, height=1, font=('Arial', 10, 'bold'), bg=BG, command=open)
upfile1_btn.place(x=550, y=450)

savefile1_btn = Button(frame1, text='Lưu vào file', width=15, height=1, font=('Arial', 10, 'bold'), bg=BG, command=save)
savefile1_btn.place(x=1420, y=450)

share_key_btn = Button(frame1, text='Chia sẻ khóa', width=30, height=2, font=('Arial', 15, 'bold'), bg=BG, command=share_key)
share_key_btn.place(x=0, y=500)

encrypt_btn = Button(frame1, text='Mã hóa', width=30, height=2, font=('Arial', 15, 'bold'), bg=BG, command=encrypt)
encrypt_btn.place(x=590, y=500)

exit1_btn = Button(frame1, text='Thoát', width=30, height=2, font=('Arial', 15, 'bold'), bg=BG, command=exit)
exit1_btn.place(x=1180, y=500)

restore_key_btn = Button(frame2, text='Khôi phục khóa', width=15, height=1, font=('Arial', 10, 'bold'), bg=BG, command=restore_key)
restore_key_btn.place(x=1400, y=50)

upfile2_btn = Button(frame2, text='Đọc từ file', width=15, height=1, font=('Arial', 10, 'bold'), bg=BG, command=decrypt_open)
upfile2_btn.place(x=550, y=480)

savefile2_btn = Button(frame2, text='Lưu vào file', width=15, height=1, font=('Arial', 10, 'bold'), bg=BG, command=decrypt_save)
savefile2_btn.place(x=1420, y=480)

decrypt_btn = Button(frame2, text='Giải mã', width=30, height=2, font=('Arial', 15, 'bold'), bg=BG, command=decrypt)
decrypt_btn.place(x=200, y=550)

exit1_btn = Button(frame2, text='Thoát', width=30, height=2, font=('Arial', 15, 'bold'), bg=BG, command=exit)
exit1_btn.place(x=1000, y=550)





















window.mainloop()
